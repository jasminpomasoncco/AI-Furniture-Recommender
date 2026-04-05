from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "productos.xlsx"

def _normalizar_registro(valor: object) -> str:
    return str(valor).strip() if valor is not None else ""


def _normalizar_columna(nombre: object) -> str:
    texto = _normalizar_registro(nombre).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _buscar_columna(columnas: dict[str, str], *alias: str) -> str | None:
    for nombre_normalizado, nombre_original in columnas.items():
        if nombre_normalizado in alias:
            return nombre_original
    return None

def cargar_productos_desde_excel(
    excel_path: Path = EXCEL_PATH,
    sheet_name: str | int = 0,
) -> list[dict[str, str]]:

    if not excel_path.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo Excel de productos: {excel_path}"
        )

    try:
        workbook = load_workbook(excel_path, data_only=True)
        if isinstance(sheet_name, int):
            worksheet = workbook[workbook.sheetnames[sheet_name]]
        else:
            worksheet = workbook[sheet_name]
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {excel_path}") from exc

    encabezados = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    columnas = {
        _normalizar_columna(encabezado): idx + 1
        for idx, encabezado in enumerate(encabezados)
        if encabezado is not None and _normalizar_columna(encabezado)
    }

    nombre_col = _buscar_columna(
        columnas,
        "nombre del producto",
        "nombre producto",
        "nombre",
    )
    descripcion_col = _buscar_columna(
        columnas,
        "descripcion comercial copywriting para ia",
        "descripcion comercial",
        "descripcion",
    )

    if nombre_col is None or descripcion_col is None:
        raise ValueError(
            "El Excel debe contener al menos 'Nombre del Producto' y 'Descripción Comercial (Copywriting para IA)'."
        )

    categoria_col = _buscar_columna(columnas, "categoria")
    medidas_col = _buscar_columna(columnas, "medidas")
    especificaciones_col = _buscar_columna(
        columnas,
        "especificaciones tecnicas",
        "especificaciones",
    )
    precios_col = _buscar_columna(columnas, "precios", "precio")

    productos: list[dict[str, str]] = []
    for row_idx in range(2, worksheet.max_row + 1):
        def celda(columna: int | None) -> str:
            if columna is None:
                return ""
            return _normalizar_registro(worksheet.cell(row_idx, columna).value)

        nombre = celda(nombre_col)
        descripcion = celda(descripcion_col)
        categoria = celda(categoria_col)
        medidas = celda(medidas_col)
        especificaciones = celda(especificaciones_col)
        precios = celda(precios_col) if precios_col else "A cotización"

        if not nombre or not descripcion:
            continue

        productos.append(
            {
                "categoria": categoria,
                "nombre": nombre,
                "descripcion": descripcion,
                "medidas": medidas,
                "especificaciones_tecnicas": especificaciones,
                "precios": precios,
            }
        )

    if not productos:
        raise ValueError("El Excel no contiene filas válidas de productos.")

    return productos

PRODUCTOS: list[dict[str, str]] = cargar_productos_desde_excel()